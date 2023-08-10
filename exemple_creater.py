# import openpyxl
#
# # Open the source Excel file
# source_file = '!Кредитна_Заявка1.27_20_10_21 (6).xlsm'
# source_workbook = openpyxl.load_workbook(source_file)
#
# # Specify the source sheet name
# source_sheet_name = 'Поручитель_1'
# source_sheet = source_workbook[source_sheet_name]
#
# # Specify the destination Excel file
# destination_file = 'example321.xlsx'
#
# # Create the destination workbook if it doesn't exist, otherwise open it
# try:
#     destination_workbook = openpyxl.load_workbook(destination_file)
# except FileNotFoundError:
#     destination_workbook = openpyxl.Workbook()
#
# # Create or use the destination sheet (named 'DestinationSheet') in the destination workbook
# destination_sheet_name = 'DestinationSheet'
# if destination_sheet_name in destination_workbook.sheetnames:
#     destination_sheet = destination_workbook[destination_sheet_name]
# else:
#     destination_sheet = destination_workbook.create_sheet(destination_sheet_name)
#
# # Copy data from source sheet to destination sheet
# for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
#     destination_sheet.append(row)
#
# # Save the changes to the destination Excel file
# destination_workbook.save(destination_file)
#
# # Close both workbooks
# source_workbook.close()
# destination_workbook.close()
import openpyxl

# Open the source Excel file
source_file = '!Кредитна_Заявка1.27_20_10_21 (6).xlsm'
source_workbook = openpyxl.load_workbook(source_file)

# Specify the source sheet name
source_sheet_name = 'Поручитель_1'
source_sheet = source_workbook[source_sheet_name]

# Specify the destination Excel file
destination_file = 'example321.xlsx'

# Create the destination workbook if it doesn't exist, otherwise open it
try:
    destination_workbook = openpyxl.load_workbook(destination_file)
except FileNotFoundError:
    destination_workbook = openpyxl.Workbook()

# Create or use the destination sheet (named 'DestinationSheet') in the destination workbook
destination_sheet_name = 'DestinationSheet'
if destination_sheet_name in destination_workbook.sheetnames:
    destination_sheet = destination_workbook[destination_sheet_name]
else:
    destination_sheet = destination_workbook.create_sheet(destination_sheet_name)

# Copy data from source sheet to destination sheet
for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
    destination_sheet.append(row)

# Delete the first sheet from the destination workbook
if 'Sheet' in destination_workbook.sheetnames:
    del destination_workbook['Sheet']

# Save the changes to the destination Excel file
destination_workbook.save(destination_file)

# Close both workbooks
source_workbook.close()
destination_workbook.close()



