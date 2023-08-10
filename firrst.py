# import openpyxl
# # wb = openpyxl.reader.excel.load_workbook(filename='qwe1.xlsx', keep_links=True, data_only=True)
# # print(wb.sheetnames)
# # wb.active = 0
# # sheet = wb.active
# # # print(sheet['A1'].value)
# # for i in range(1,12):
# #     print(sheet['A'+str(i)].value, sheet['B'+str(i)].value, sheet['C'+str(i)].value)
# wb = openpyxl.reader.excel.load_workbook(filename='!Кредитна_Заявка1.27_20_10_21 (4).xlsm',read_only=True,keep_links =True,data_only=True, rich_text=True)
# # print(wb.sheetnames)
# wb.active = 1
# sheet = wb.active
# sheet.read()

# import openpyxl
# import pandas as pd
#
# exel_file = '!Кредитна_Заявка1.27_20_10_21.xlsm'
# sheet_name = 'Позичальник'
#
# df = pd.read_excel(exel_file,sheet_name)
#
# print(df.head(2))



# import openpyxl
#
# # Open the source Excel file
# source_file = '!Кредитна_Заявка1.27_20_10_21.xlsm'
# source_workbook = openpyxl.load_workbook(source_file)
# source_sheet = source_workbook.active  # Assuming you want to use the active sheet
#
# # Open the destination Excel file
# destination_file = 'destination.xlsx'
# destination_workbook = openpyxl.load_workbook(destination_file)
# destination_sheet = destination_workbook.active  # Assuming you want to use the active sheet
#
# # Copy data from source sheet to destination sheet
# for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
#     destination_sheet.append(row)
#
# # Save the changes to the destination Excel file
# destination_workbook.save(destination_file)
#
# # Close both workbooks
# source_workbook.close()
# destination_workbook.close()



# import openpyxl
#
# # Open the source Excel file
# source_file = '!Кредитна_Заявка1.27_20_10_21.xlsm'
# source_workbook = openpyxl.load_workbook(source_file)
#
# # Specify the source sheet name
# source_sheet_name = 'Позичальник'  # Replace with the actual source sheet name
# source_sheet = source_workbook[source_sheet_name]
#
# # Open the destination Excel file
# destination_file = 'destination.xlsx'
# destination_workbook = openpyxl.load_workbook(destination_file)
# destination_sheet = destination_workbook.active  # Assuming you want to use the active sheet
#
# # Copy data from source sheet to destination sheet
# for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
#     destination_sheet.append(row)
#
# # Save the changes to the destination Excel file
# destination_workbook.save(destination_file)
#
# # Close both workbooks
# source_workbook.close()
# destination_workbook.close()

# import openpyxl
# import os
#
# from openpyxl.styles import NamedStyle
#
# # Specify the source Excel file
# source_file = '!Кредитна_Заявка1.27_20_10_21 (6).xlsm'
#
# # Specify the source sheet name
# source_sheet_name = 'Позичальник'  # Replace with the actual source sheet name
#
# # Specify the destination Excel file (will be created if not exists)
# destination_file = 'destination.xlsx'
#
# # Close the destination file if it's open
# try:
#     os.remove(destination_file)
# except FileNotFoundError:
#     pass
#
# # Open the source Excel file
# source_workbook = openpyxl.load_workbook(source_file)
# source_sheet = source_workbook[source_sheet_name]
#
# # Extract data from cells B5, C5, and D5 in the source sheet
# data_to_extract = [source_sheet.cell(row=5, column=2).value,  # B5
#                    source_sheet.cell(row=5, column=3).value,  # C5
#                    source_sheet.cell(row=5, column=4).value]  # D5
#
# # Extract data from cell B6 in the source sheet
# data_to_append_a2 = source_sheet.cell(row=6, column=2).value  # B6
# data_to_append_b2 = source_sheet.cell(row=7, column=2).value  # B7
# data_to_append_a3 = source_sheet.cell(row=6, column=8).value  # H6
# data_to_append_b3 = source_sheet.cell(row=7, column=8).value  # H7
# data_to_append_b3 = source_sheet.cell(row=7, column=8).value  # H7
# data_to_append_b3 = source_sheet.cell(row=7, column=8).value  # H7
#
# data_to_append_n6 = source_sheet.cell(row=6, column=14).value  # N6
# data_to_append_o6 = source_sheet.cell(row=6, column=15).value  # O6
#
# data_to_append_n7 = source_sheet.cell(row=7, column=14).value  # N6
# data_to_append_o7 = source_sheet.cell(row=7, column=15).value  # O7
#
# data_to_append_y6 = source_sheet.cell(row=6, column=25).value  # N6
# data_to_append_y7 = source_sheet.cell(row=7, column=26).value  # O7
#
# data_to_append_b5 = source_sheet.cell(row=7, column=25).value  # H7
#
# # Create or open the destination Excel file
# destination_workbook = openpyxl.Workbook()  # Create a new workbook
# destination_sheet = destination_workbook.active
#
#
#
# # Specify the destination cell where you want to paste the text
# for col_idx, value in enumerate(data_to_extract, start=1):
#     destination_sheet.cell(row=1, column=col_idx, value=value)
#
# # Adjust column width for column B to display the date correctly
# destination_sheet.column_dimensions['B'].width = 15  # Adjust width for column B
# destination_sheet.column_dimensions['A'].width = 15  # Adjust width for column B
#
# # Format cell B5 as a date cell
# date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')
# destination_sheet['B5'].style = date_style
#
# # Append the data from B6 to cell A2 in the destination sheet
# destination_sheet['A2'] = data_to_append_a2
# destination_sheet['B2'] = data_to_append_b2
# destination_sheet['A3'] = data_to_append_a3
# destination_sheet['B3'] = data_to_append_b3
# destination_sheet['A4'] = f"{data_to_append_n6 or ''} {data_to_append_o6 or ''}"
# destination_sheet['B4'] = f"{data_to_append_n7 or ''} {data_to_append_o7 or ''}"
# destination_sheet['A5'] = f"{data_to_append_y6 or ''} {data_to_append_y7 or ''}"
# destination_sheet['B5'] = data_to_append_b5
#
# # Save the destination workbook
# destination_workbook.save(destination_file)
#
# # Close both workbooks
# source_workbook.close()
# destination_workbook.close()



# zxc
# import openpyxl
# from openpyxl.styles import NamedStyle, Alignment
# import warnings
#
# # Ignore the specific Data Validation warning
# warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")
#
# # Specify the source Excel file
# source_file = '!Кредитна_Заявка1.27_20_10_21 (6).xlsm'
#
# # Specify the source sheet name
# source_sheet_name = 'Позичальник'
#
# # Specify the destination Excel file
# destination_file = 'destination.xlsx'
#
# # Open the source Excel file
# source_workbook = openpyxl.load_workbook(source_file)
# source_sheet = source_workbook[source_sheet_name]
#
# # Create the destination Excel file
# destination_workbook = openpyxl.Workbook()
# destination_sheet = destination_workbook.active
# #
# # Adjust column widths
# for col_letter in ['B']:
#     destination_sheet.column_dimensions[col_letter].width = 15
# destination_sheet.column_dimensions['A'].width = 45
#
# # Define the alignment style for left alignment
# left_alignment = Alignment(horizontal='left')
#
# data_to_append = {
#     'A1': f"{source_sheet.cell(row=5, column=2).value or ''} ",
#     'A2': source_sheet.cell(row=6, column=2).value,
#     'B2': source_sheet.cell(row=7, column=2).value,
#     'A3': source_sheet.cell(row=6, column=8).value,
#     'B3': source_sheet.cell(row=7, column=8).value,
#     'A4': f"{source_sheet.cell(row=6, column=14).value or ''} {source_sheet.cell(row=6, column=15).value or ''}",
#     'B4': f"{source_sheet.cell(row=7, column=14).value or ''} {source_sheet.cell(row=7, column=15).value or ''}",
#     'A5': f"{source_sheet.cell(row=6, column=25).value or ''} {source_sheet.cell(row=7, column=26).value or ''}",
#     'B5': source_sheet.cell(row=7, column=25).value,
#     'A6': source_sheet.cell(row=6,column=29).value,
#     'B6': source_sheet.cell(row=7,column=29).value,
#     'A7':source_sheet.cell(row=9,column=2).value,
#     'B7':source_sheet.cell(row=10,column=2).value,
#     'A8':source_sheet.cell(row=9,column=7).value,
#     'B8':source_sheet.cell(row=10,column=7).value,
#     'A9':source_sheet.cell(row=9,column=28).value,
#     'B9': source_sheet.cell(row=10, column=42).value,
#
#
#
# }
#
# for cell, value in data_to_append.items():
#     destination_sheet[cell] = value
#     if cell.startswith('B'):
#         destination_sheet[cell].alignment = left_alignment  # Apply left alignment to cell in column B
#
# # Format cell B5 as a date cell
# date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')
# destination_sheet['B5'].style = date_style
#
# # Save the destination workbook
# destination_workbook.save(destination_file)
#
# # Close both workbooks
# source_workbook.close()
# destination_workbook.close()
# print('finished')
# pip install -r requirements.txt


import openpyxl
from openpyxl.styles import NamedStyle, Alignment
import warnings
import re
import os
import shutil
import time

# Start the timer
start_time = time.time()

# Ignore the specific Data Validation warning
warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")

source_folder = 'main'
output_folder = 'new'

# Recreate the output folder (delete and create it again)
if os.path.exists(output_folder):
    shutil.rmtree(output_folder)
os.mkdir(output_folder)

for source_file in os.listdir(source_folder):
    if source_file.endswith('.xlsm'):
        source_file_path = os.path.join(source_folder, source_file)
        # Specify the source Excel file
        # source_file = '!Кредитна_Заявка1.27_20_10_21 (6).xlsm'

        # Specify the source sheet name
        source_sheet_name1 = 'Позичальник'
        source_sheet_name2 = 'Поручитель_1'

        # Specify the destination Excel file
        # destination_file = 'destination.xlsx'

        # Open the source Excel file
        source_workbook = openpyxl.load_workbook(source_file_path)

        source_sheet = source_workbook[source_sheet_name1]
        file_name_value = str(source_sheet.cell(row=7, column=29).value)
        source_sheet2 = source_workbook[source_sheet_name2]

        # Create the destination Excel file
        destination_workbook = openpyxl.Workbook()
        destination_sheet = destination_workbook.active

        # Adjust column widths
        for col_letter in ['B','H']:
            destination_sheet.column_dimensions[col_letter].width = 20
        destination_sheet.column_dimensions['A'].width = 45
        destination_sheet.column_dimensions['G'].width = 45

        # Define the alignment style for left alignment
        left_alignment = Alignment(horizontal='left')

        data_to_append1 = {
            'A1': f"{source_sheet.cell(row=5, column=2).value or ''} ",
            'A2': source_sheet.cell(row=6, column=2).value,
            'B2': source_sheet.cell(row=7, column=2).value,
            'A3': source_sheet.cell(row=6, column=8).value,
            'B3': source_sheet.cell(row=7, column=8).value,
            'A4': f"{source_sheet.cell(row=6, column=14).value or ''} {source_sheet.cell(row=6, column=15).value or ''}",
            'B4': f"{source_sheet.cell(row=7, column=14).value or ''} {source_sheet.cell(row=7, column=15).value or ''}",
            'A5': f"{source_sheet.cell(row=6, column=25).value or ''} {source_sheet.cell(row=7, column=26).value or ''}",
            'B5': source_sheet.cell(row=7, column=25).value,
            'A6': source_sheet.cell(row=6, column=29).value,
            'B6': source_sheet.cell(row=7, column=29).value,
            'A7': source_sheet.cell(row=9, column=2).value,
            'B7': source_sheet.cell(row=10, column=2).value,
            'A8': source_sheet.cell(row=9, column=7).value,
            'B8': source_sheet.cell(row=10, column=7).value,
            'A9': source_sheet.cell(row=9, column=28).value,
            'B9': source_sheet.cell(row=10, column=42).value,
            'A11': source_sheet.cell(row=12, column=2).value,
            'A12': source_sheet.cell(row=13, column=2).value,
            'B12': source_sheet.cell(row=14, column=2).value,
            'A13': source_sheet.cell(row=13, column=8).value,
            'B13': source_sheet.cell(row=14, column=8).value,
            'A14': source_sheet.cell(row=13, column=11).value,
            'B14': source_sheet.cell(row=14, column=11).value,
            'A15': source_sheet.cell(row=13, column=15).value,
            'B15': source_sheet.cell(row=14, column=15).value,
            'A16': source_sheet.cell(row=13, column=19).value,
            'B16': source_sheet.cell(row=14, column=19).value,
            'A18': source_sheet.cell(row=16, column=2).value,
            'A19': source_sheet.cell(row=17, column=2).value,
            'B19': source_sheet.cell(row=18, column=2).value,
            'A20': source_sheet.cell(row=17, column=5).value,
            'B20': source_sheet.cell(row=18, column=5).value,
            'A21': source_sheet.cell(row=17, column=11).value,
            'B21': source_sheet.cell(row=18, column=11).value,
            'A22': source_sheet.cell(row=17, column=17).value,
            'B22': source_sheet.cell(row=18, column=17).value,
            'A23': source_sheet.cell(row=17, column=25).value,
            'B23': source_sheet.cell(row=18, column=25).value,
            'A24': source_sheet.cell(row=17, column=32).value,
            'B24': source_sheet.cell(row=18, column=32).value,
            'A25': source_sheet.cell(row=17, column=35).value,
            'B25': source_sheet.cell(row=18, column=35).value,
            'A26': source_sheet.cell(row=17, column=38).value,
            'B26': source_sheet.cell(row=18, column=38).value,
            'A28': source_sheet.cell(row=20, column=2).value,
            'A29': source_sheet.cell(row=21, column=2).value,
            'B29': source_sheet.cell(row=22, column=2).value,
            'A30': source_sheet.cell(row=21, column=5).value,
            'B30': source_sheet.cell(row=22, column=5).value,
            'A31': source_sheet.cell(row=21, column=11).value,
            'B31': source_sheet.cell(row=22, column=11).value,
            'A32': source_sheet.cell(row=21, column=17).value,
            'B32': source_sheet.cell(row=22, column=17).value,
            'A33': source_sheet.cell(row=21, column=25).value,
            'B33': source_sheet.cell(row=22, column=25).value,
            'A34': source_sheet.cell(row=21, column=32).value,
            'B34': source_sheet.cell(row=22, column=32).value,
            'A35': source_sheet.cell(row=21, column=35).value,
            'B35': source_sheet.cell(row=22, column=35).value,
            'A36': source_sheet.cell(row=21, column=38).value,
            'B36': source_sheet.cell(row=22, column=38).value,
            'A37': source_sheet.cell(row=21, column=41).value,
            'B37': source_sheet.cell(row=22, column=41).value,
            'A39': source_sheet.cell(row=28, column=2).value,
            'A40': source_sheet.cell(row=29, column=20).value,
            'B40': source_sheet.cell(row=30, column=20).value,
            'A41': source_sheet.cell(row=29, column=27).value,
            'B41': source_sheet.cell(row=30, column=27).value,
            'A43': source_sheet.cell(row=54, column=2).value,
            'A44': source_sheet.cell(row=55, column=2).value,
            'B44': source_sheet.cell(row=56, column=2).value,
            'A45': source_sheet.cell(row=55, column=8).value,
            'B45': source_sheet.cell(row=56, column=8).value,
            'A46': source_sheet.cell(row=55, column=14).value,
            'B46': source_sheet.cell(row=56, column=14).value,
            'A47': source_sheet.cell(row=55, column=25).value,
            'B47': source_sheet.cell(row=56, column=25).value,
            'A48': source_sheet.cell(row=55, column=29).value,
            'B48': source_sheet.cell(row=56, column=29).value,
            'A50': source_sheet.cell(row=58, column=2).value,
            'A51': source_sheet.cell(row=59, column=2).value,
            'B51': source_sheet.cell(row=60, column=2).value,
            'A52': source_sheet.cell(row=59, column=8).value,
            'B52': source_sheet.cell(row=60, column=8).value,
            'A53': source_sheet.cell(row=59, column=11).value,
            'B53': source_sheet.cell(row=60, column=11).value,
            'A54': source_sheet.cell(row=59, column=15).value,
            'B54': source_sheet.cell(row=60, column=15).value,
            'A55': source_sheet.cell(row=59, column=19).value,
            'B55': source_sheet.cell(row=60, column=19).value,
            'A56': source_sheet.cell(row=59, column=37).value,
            'B56': source_sheet.cell(row=60, column=37).value,
            'A58': source_sheet.cell(row=83, column=2).value,
            'A59': source_sheet.cell(row=85, column=2).value,
            'B59': source_sheet.cell(row=86, column=2).value,
            'A60': source_sheet.cell(row=85, column=30).value,
            'B60': source_sheet.cell(row=86, column=30).value,
            'A62': source_sheet.cell(row=88, column=2).value,
            'A63': source_sheet.cell(row=89, column=2).value,
            'B63': source_sheet.cell(row=90, column=2).value,
            'A64': source_sheet.cell(row=89, column=9).value,
            'B64': source_sheet.cell(row=90, column=9).value,
            'A65': source_sheet.cell(row=89, column=14).value,
            'B65': source_sheet.cell(row=90, column=14).value,
            'A66': source_sheet.cell(row=89, column=21).value,
            'B66': source_sheet.cell(row=90, column=21).value,
            'A67': source_sheet.cell(row=89, column=43).value,
            'B67': source_sheet.cell(row=90, column=43).value,


        }

        data_to_append2 = {
            'G1': f"{source_sheet2.cell(row=5, column=2).value or ''} ",
            'G2': source_sheet2.cell(row=6,column=2).value,
            'H2': source_sheet2.cell(row=7,column=2).value,
            'G3': source_sheet2.cell(row=6, column=8).value,
            'H3': source_sheet2.cell(row=7, column=8).value,
            'G4': source_sheet2.cell(row=6, column=14).value,
            'H4': source_sheet2.cell(row=7, column=14).value,
            'G5': source_sheet2.cell(row=6, column=24).value,
            'H5': source_sheet2.cell(row=7, column=24).value,
            'G6': source_sheet2.cell(row=6, column=28).value,
            'H6': source_sheet2.cell(row=7, column=28).value,
            'G11': source_sheet2.cell(row=12, column=2).value,
            'G12': source_sheet2.cell(row=13, column=2).value,
            'H12': source_sheet2.cell(row=14, column=2).value,
            'G13': source_sheet2.cell(row=13, column=8).value,
            'H13': source_sheet2.cell(row=14, column=8).value,
            'G14': source_sheet2.cell(row=13, column=11).value,
            'H14': source_sheet2.cell(row=14, column=11).value,
            'G15': source_sheet2.cell(row=13, column=15).value,
            'H15': source_sheet2.cell(row=14, column=15).value,
            'G16': source_sheet2.cell(row=13, column=19).value,
            'H16': source_sheet2.cell(row=14, column=19).value,





        }

        combined_data = {**data_to_append1, **data_to_append2}

        for cell, value in combined_data.items():
            destination_sheet[cell] = value
            if cell.startswith('B') or cell.startswith('H'):
                destination_sheet[cell].alignment = left_alignment  # Apply left alignment to cell in column B

        # Format cell B5 as a date cell and apply left alignment
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY', alignment=left_alignment)
        destination_sheet['B5'].style = date_style
        destination_sheet['H5'].style = date_style

        cleaned_file_name = re.sub(r'[<>:"/\|?*]', '', file_name_value)[:30]  # Limit to 30 characters
        destination_file = f'{cleaned_file_name}.xlsx'

        # Save the destination workbook
        destination_workbook.save(os.path.join(output_folder, destination_file))

        # Close both workbooks
        source_workbook.close()
        destination_workbook.close()
# Stop the timer
end_time = time.time()
elapsed_time = end_time - start_time
print(f"Finished in {elapsed_time:.2f} seconds")


