# import openpyxl
# from openpyxl.styles import NamedStyle, Alignment
# import warnings
# import re
# import os
# import shutil
# import time
#
#
# # Start the timer
# start_time = time.time()
#
# # Ignore the specific Data Validation warning
# warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")
#
# source_folder = 'main'
# output_folder = 'new'
#
# # Get the total number of files
# total_files = sum(1 for root, _, files in os.walk(source_folder) for file in files if file.endswith('.xlsm'))
#
# processed_files = 0
# update_interval = 60  # Update progress every 1 minute
# last_update_time = time.time()
#
# # Recreate the output folder (delete and create it again)
# if os.path.exists(output_folder):
#     shutil.rmtree(output_folder)
# os.mkdir(output_folder)
#
# # Check any exel file in folder 'main' and other folders it contains
# for root, _, files in os.walk(source_folder):
#     for source_file in files:
#         if source_file.endswith('.xlsm') or source_file.endswith('.XLSM'):
#             source_file_path = os.path.join(root, source_file)
#
#             # Specify the source sheet name
#             source_sheet_name1 = 'Позичальник'
#             source_sheet_name2 = 'Поручитель_1'
#
#             if source_sheet_name1 not in source_workbook.sheetnames:
#                 print(f"Worksheet '{source_sheet_name1}' does not exist in {source_file}. Skipping.")
#                 source_workbook.close()
#                 continue
#
#             # Check if the second sheet exists and if not, move to the next file
#             if source_sheet_name2 not in source_workbook.sheetnames:
#                 print(f"Worksheet '{source_sheet_name2}' does not exist in {source_file}. Skipping.")
#                 source_workbook.close()
#                 continue
#
#             # Open the source Excel file
#             source_workbook = openpyxl.load_workbook(source_file_path)
#             source_sheet = source_workbook[source_sheet_name1]
#             source_sheet2 = source_workbook[source_sheet_name2]
#
#             file_name_value = str(source_sheet.cell(row=7, column=29).value)
#
#             # Create the destination Excel file
#             destination_workbook = openpyxl.Workbook()
#             destination_sheet = destination_workbook.active
#
#             # Adjust column widths
#             for col_letter in ['B','H']:
#                 destination_sheet.column_dimensions[col_letter].width = 20
#             destination_sheet.column_dimensions['A'].width = 45
#             destination_sheet.column_dimensions['G'].width = 45
#
#             # Define the alignment style for left alignment
#             left_alignment = Alignment(horizontal='left')
#
#             data_to_append1 = {
#                 'A1': f"ОСОБА 1",
#                 'A2': source_sheet.cell(row=6, column=2).value,
#                 'B2': source_sheet.cell(row=7, column=2).value,
#                 'A3': source_sheet.cell(row=6, column=8).value,
#                 'B3': source_sheet.cell(row=7, column=8).value,
#                 'A4': f"{source_sheet.cell(row=6, column=14).value or ''} {source_sheet.cell(row=6, column=15).value or ''}",
#                 'B4': f"{source_sheet.cell(row=7, column=14).value or ''} {source_sheet.cell(row=7, column=15).value or ''}",
#                 'A5': f"{source_sheet.cell(row=6, column=25).value or ''} {source_sheet.cell(row=7, column=26).value or ''}",
#                 'B5': source_sheet.cell(row=7, column=25).value,
#                 'A6': source_sheet.cell(row=6, column=29).value,
#                 'B6': source_sheet.cell(row=7, column=29).value,
#                 # 'A7': source_sheet.cell(row=9, column=2).value,
#                 # 'B7': source_sheet.cell(row=10, column=2).value,
#                 'A7': source_sheet.cell(row=9, column=7).value,
#                 'B7': source_sheet.cell(row=10, column=7).value,
#                 # 'A9': source_sheet.cell(row=9, column=28).value,
#                 # 'B9': source_sheet.cell(row=10, column=42).value,
#                 'A9': source_sheet.cell(row=12, column=2).value,
#                 'A10': source_sheet.cell(row=13, column=2).value,
#                 'B10': source_sheet.cell(row=14, column=2).value,
#                 'A11': source_sheet.cell(row=13, column=8).value,
#                 'B11': source_sheet.cell(row=14, column=8).value,
#                 'A12': source_sheet.cell(row=13, column=11).value,
#                 'B12': source_sheet.cell(row=14, column=11).value,
#                 'A13': source_sheet.cell(row=13, column=15).value,
#                 'B13': source_sheet.cell(row=14, column=15).value,
#                 'A14': source_sheet.cell(row=13, column=19).value,
#                 'B14': source_sheet.cell(row=14, column=19).value,
#                 'A16': source_sheet.cell(row=16, column=2).value,
#                 # 'A19': source_sheet.cell(row=17, column=2).value,
#                 # 'B19': source_sheet.cell(row=18, column=2).value,
#                 'A17': source_sheet.cell(row=17, column=5).value,
#                 'B17': source_sheet.cell(row=18, column=5).value,
#                 'A18': source_sheet.cell(row=17, column=11).value,
#                 'B18': source_sheet.cell(row=18, column=11).value,
#                 'A19': source_sheet.cell(row=17, column=17).value,
#                 'B19': source_sheet.cell(row=18, column=17).value,
#                 'A20': source_sheet.cell(row=17, column=25).value,
#                 'B20': source_sheet.cell(row=18, column=25).value,
#                 'A21': source_sheet.cell(row=17, column=32).value,
#                 'B21': source_sheet.cell(row=18, column=32).value,
#                 'A22': source_sheet.cell(row=17, column=35).value,
#                 'B22': source_sheet.cell(row=18, column=35).value,
#                 'A23': source_sheet.cell(row=17, column=38).value,
#                 'B23': source_sheet.cell(row=18, column=38).value,
#                 'A25': source_sheet.cell(row=20, column=2).value,
#                 # 'A29': source_sheet.cell(row=21, column=2).value,
#                 # 'B29': source_sheet.cell(row=22, column=2).value,
#                 'A26': source_sheet.cell(row=21, column=5).value,
#                 'B26': source_sheet.cell(row=22, column=5).value,
#                 'A27': source_sheet.cell(row=21, column=11).value,
#                 'B27': source_sheet.cell(row=22, column=11).value,
#                 'A28': source_sheet.cell(row=21, column=17).value,
#                 'B28': source_sheet.cell(row=22, column=17).value,
#                 'A29': source_sheet.cell(row=21, column=25).value,
#                 'B29': source_sheet.cell(row=22, column=25).value,
#                 'A30': source_sheet.cell(row=21, column=32).value,
#                 'B30': source_sheet.cell(row=22, column=32).value,
#                 'A31': source_sheet.cell(row=21, column=35).value,
#                 'B31': source_sheet.cell(row=22, column=35).value,
#                 'A32': source_sheet.cell(row=21, column=38).value,
#                 'B32': source_sheet.cell(row=22, column=38).value,
#                 'A33': source_sheet.cell(row=21, column=41).value,
#                 'B33': source_sheet.cell(row=22, column=41).value,
#                 'A35': source_sheet.cell(row=28, column=2).value,
#                 'A36': source_sheet.cell(row=29, column=20).value,
#                 'B36': source_sheet.cell(row=30, column=20).value,
#                 'A37': source_sheet.cell(row=29, column=27).value,
#                 'B37': source_sheet.cell(row=30, column=27).value,
#                 'A39': source_sheet.cell(row=54, column=2).value,
#                 'A40': source_sheet.cell(row=55, column=2).value,
#                 'B40': source_sheet.cell(row=56, column=2).value,
#                 'A41': source_sheet.cell(row=55, column=8).value,
#                 'B41': source_sheet.cell(row=56, column=8).value,
#                 'A42': source_sheet.cell(row=55, column=14).value,
#                 'B42': source_sheet.cell(row=56, column=14).value,
#                 'A43': source_sheet.cell(row=55, column=25).value,
#                 'B43': source_sheet.cell(row=56, column=25).value,
#                 'A44': source_sheet.cell(row=55, column=29).value,
#                 'B44': source_sheet.cell(row=56, column=29).value,
#                 'A46': source_sheet.cell(row=58, column=2).value,
#                 'A47': source_sheet.cell(row=59, column=2).value,
#                 'B47': source_sheet.cell(row=60, column=2).value,
#                 'A48': source_sheet.cell(row=59, column=8).value,
#                 'B48': source_sheet.cell(row=60, column=8).value,
#                 'A49': source_sheet.cell(row=59, column=11).value,
#                 'B49': source_sheet.cell(row=60, column=11).value,
#                 'A50': source_sheet.cell(row=59, column=15).value,
#                 'B50': source_sheet.cell(row=60, column=15).value,
#                 'A51': source_sheet.cell(row=59, column=19).value,
#                 'B51': source_sheet.cell(row=60, column=19).value,
#                 'A52': source_sheet.cell(row=59, column=37).value,
#                 'B52': source_sheet.cell(row=60, column=37).value,
#                 # 'A58': source_sheet.cell(row=83, column=2).value,
#                 # 'A59': source_sheet.cell(row=85, column=2).value,
#                 # 'B59': source_sheet.cell(row=86, column=2).value,
#                 # 'A60': source_sheet.cell(row=85, column=30).value,
#                 # 'B60': source_sheet.cell(row=86, column=30).value,
#                 'A54': source_sheet.cell(row=88, column=2).value,
#                 'A55': source_sheet.cell(row=89, column=2).value,
#                 'B55': source_sheet.cell(row=90, column=2).value,
#                 'A56': source_sheet.cell(row=89, column=9).value,
#                 'B56': source_sheet.cell(row=90, column=9).value,
#                 'A57': source_sheet.cell(row=89, column=14).value,
#                 'B57': source_sheet.cell(row=90, column=14).value,
#                 'A58': source_sheet.cell(row=89, column=21).value,
#                 'B58': source_sheet.cell(row=90, column=21).value,
#                 'A59': source_sheet.cell(row=89, column=43).value,
#                 'B59': source_sheet.cell(row=90, column=43).value,
#             }
#
#             data_to_append2 = {
#                 'G1': f"ОСОБА 2",
#                 'G2': source_sheet2.cell(row=6,column=2).value,
#                 'H2': source_sheet2.cell(row=7,column=2).value,
#                 'G3': source_sheet2.cell(row=6, column=8).value,
#                 'H3': source_sheet2.cell(row=7, column=8).value,
#                 'G4': source_sheet2.cell(row=6, column=14).value,
#                 'H4': source_sheet2.cell(row=7, column=14).value,
#                 'G5': source_sheet2.cell(row=6, column=24).value,
#                 'H5': source_sheet2.cell(row=7, column=24).value,
#                 'G6': source_sheet2.cell(row=6, column=28).value,
#                 'H6': source_sheet2.cell(row=7, column=28).value,
#                 'G9': source_sheet2.cell(row=12, column=2).value,
#                 'G10': source_sheet2.cell(row=13, column=2).value,
#                 'H10': source_sheet2.cell(row=14, column=2).value,
#                 'G11': source_sheet2.cell(row=13, column=8).value,
#                 'H11': source_sheet2.cell(row=14, column=8).value,
#                 'G12': source_sheet2.cell(row=13, column=11).value,
#                 'H12': source_sheet2.cell(row=14, column=11).value,
#                 'G13': source_sheet2.cell(row=13, column=15).value,
#                 'H13': source_sheet2.cell(row=14, column=15).value,
#                 'G14': source_sheet2.cell(row=13, column=19).value,
#                 'H14': source_sheet2.cell(row=14, column=19).value,
#             }
#
#             combined_data = {**data_to_append1, **data_to_append2}
#
#             for cell, value in combined_data.items():
#                 destination_sheet[cell] = value
#                 if cell.startswith('B') or cell.startswith('H'):
#                     destination_sheet[cell].alignment = left_alignment  # Apply left alignment to cell in column B
#
#             # Format cell B5 as a date cell and apply left alignment
#             date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY', alignment=left_alignment)
#             destination_sheet['B5'].style = date_style
#             destination_sheet['H5'].style = date_style
#
#             cleaned_file_name = re.sub(r'[<>:"/\|?*]', '', file_name_value)[:30]  # Limit to 30 characters
#             destination_file = f'{cleaned_file_name}.xlsx'
#
#             # Save the destination workbook
#             destination_workbook.save(os.path.join(output_folder, destination_file))
#
#             # Close both workbooks
#             source_workbook.close()
#             destination_workbook.close()
#
#             processed_files += 1
#
#             current_time = time.time()
#             if current_time - last_update_time >= update_interval:
#                 progress = (processed_files / total_files) * 100
#                 print(f"Progress: {progress:.2f}% ({processed_files}/{total_files} files)")
#                 last_update_time = current_time
# # Stop the timer
# end_time = time.time()
# elapsed_time = end_time - start_time
# print(f"Finished in {elapsed_time:.2f} seconds")

###############################################################################

import openpyxl
from openpyxl.styles import NamedStyle, Alignment
import warnings
import re
import os
import shutil
import time

# Start the timer
start_time = time.time()

# Ignore the specific Data Validation warning
warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")

source_folder = 'main'
output_folder = 'new'

# Get the total number of files
total_files = sum(1 for root, _, files in os.walk(source_folder) for file in files if file.endswith('.xlsm'))

processed_files = 0
update_interval = 60  # Update progress every 1 minute
last_update_time = time.time()

# Recreate the output folder (delete and create it again)
if os.path.exists(output_folder):
    shutil.rmtree(output_folder)
os.mkdir(output_folder)

# Define the source sheet names
source_sheet_name1 = 'Позичальник'
source_sheet_name2 = 'Поручитель_1'

# Iterate through all Excel files in the source folder and its subdirectories
for root, _, files in os.walk(source_folder):
    for source_file in files:
        if source_file.endswith('.xlsm') or source_file.endswith('.XLSM'):
            source_file_path = os.path.join(root, source_file)

            try:
                # Open the source Excel file
                source_workbook = openpyxl.load_workbook(source_file_path, data_only=True, read_only=True)

                # Check if the first sheet exists and if not, move to the next file
                if source_sheet_name1 not in source_workbook.sheetnames:
                    print(f"Worksheet '{source_sheet_name1}' does not exist in {source_file}. Skipping.")
                    source_workbook.close()
                    continue

                # Check if the second sheet exists and if not, move to the next file
                if source_sheet_name2 not in source_workbook.sheetnames:
                    print(f"Worksheet '{source_sheet_name2}' does not exist in {source_file}. Skipping.")
                    source_workbook.close()
                    continue

                source_sheet = source_workbook[source_sheet_name1]
                source_sheet2 = source_workbook[source_sheet_name2]

                file_name_value = str(source_sheet.cell(row=7, column=29).value)

                # Create the destination Excel file
                destination_workbook = openpyxl.Workbook()
                destination_sheet = destination_workbook.active

                # Adjust column widths
                for col_letter in ['B', 'H']:
                    destination_sheet.column_dimensions[col_letter].width = 20
                destination_sheet.column_dimensions['A'].width = 45
                destination_sheet.column_dimensions['G'].width = 45

                # Define the alignment style for left alignment
                left_alignment = Alignment(horizontal='left')

                data_to_append1 = {
                    'A1': f"ОСОБА 1",
                    'A2': source_sheet.cell(row=6, column=2).value,
                    'B2': source_sheet.cell(row=7, column=2).value,
                    'A3': source_sheet.cell(row=6, column=8).value,
                    'B3': source_sheet.cell(row=7, column=8).value,
                    'A4': f"{source_sheet.cell(row=6, column=14).value or ''} {source_sheet.cell(row=6, column=15).value or ''}",
                    'B4': f"{source_sheet.cell(row=7, column=14).value or ''} {source_sheet.cell(row=7, column=15).value or ''}",
                    'A5': f"{source_sheet.cell(row=6, column=25).value or ''} {source_sheet.cell(row=7, column=26).value or ''}",
                    'B5': source_sheet.cell(row=7, column=25).value,
                    'A6': source_sheet.cell(row=6, column=29).value,
                    'B6': source_sheet.cell(row=7, column=29).value,
                    # 'A7': source_sheet.cell(row=9, column=2).value,
                    # 'B7': source_sheet.cell(row=10, column=2).value,
                    'A7': source_sheet.cell(row=9, column=7).value,
                    'B7': source_sheet.cell(row=10, column=7).value,
                    # 'A9': source_sheet.cell(row=9, column=28).value,
                    # 'B9': source_sheet.cell(row=10, column=42).value,
                    'A9': source_sheet.cell(row=12, column=2).value,
                    'A10': source_sheet.cell(row=13, column=2).value,
                    'B10': source_sheet.cell(row=14, column=2).value,
                    'A11': source_sheet.cell(row=13, column=8).value,
                    'B11': source_sheet.cell(row=14, column=8).value,
                    'A12': source_sheet.cell(row=13, column=11).value,
                    'B12': source_sheet.cell(row=14, column=11).value,
                    'A13': source_sheet.cell(row=13, column=15).value,
                    'B13': source_sheet.cell(row=14, column=15).value,
                    'A14': source_sheet.cell(row=13, column=19).value,
                    'B14': source_sheet.cell(row=14, column=19).value,
                    'A16': source_sheet.cell(row=16, column=2).value,
                    # 'A19': source_sheet.cell(row=17, column=2).value,
                    # 'B19': source_sheet.cell(row=18, column=2).value,
                    'A17': source_sheet.cell(row=17, column=5).value,
                    'B17': source_sheet.cell(row=18, column=5).value,
                    'A18': source_sheet.cell(row=17, column=11).value,
                    'B18': source_sheet.cell(row=18, column=11).value,
                    'A19': source_sheet.cell(row=17, column=17).value,
                    'B19': source_sheet.cell(row=18, column=17).value,
                    'A20': source_sheet.cell(row=17, column=25).value,
                    'B20': source_sheet.cell(row=18, column=25).value,
                    'A21': source_sheet.cell(row=17, column=32).value,
                    'B21': source_sheet.cell(row=18, column=32).value,
                    'A22': source_sheet.cell(row=17, column=35).value,
                    'B22': source_sheet.cell(row=18, column=35).value,
                    'A23': source_sheet.cell(row=17, column=38).value,
                    'B23': source_sheet.cell(row=18, column=38).value,
                    'A25': source_sheet.cell(row=20, column=2).value,
                    # 'A29': source_sheet.cell(row=21, column=2).value,
                    # 'B29': source_sheet.cell(row=22, column=2).value,
                    'A26': source_sheet.cell(row=21, column=5).value,
                    'B26': source_sheet.cell(row=22, column=5).value,
                    'A27': source_sheet.cell(row=21, column=11).value,
                    'B27': source_sheet.cell(row=22, column=11).value,
                    'A28': source_sheet.cell(row=21, column=17).value,
                    'B28': source_sheet.cell(row=22, column=17).value,
                    'A29': source_sheet.cell(row=21, column=25).value,
                    'B29': source_sheet.cell(row=22, column=25).value,
                    'A30': source_sheet.cell(row=21, column=32).value,
                    'B30': source_sheet.cell(row=22, column=32).value,
                    'A31': source_sheet.cell(row=21, column=35).value,
                    'B31': source_sheet.cell(row=22, column=35).value,
                    'A32': source_sheet.cell(row=21, column=38).value,
                    'B32': source_sheet.cell(row=22, column=38).value,
                    'A33': source_sheet.cell(row=21, column=41).value,
                    'B33': source_sheet.cell(row=22, column=41).value,
                    'A35': source_sheet.cell(row=28, column=2).value,
                    'A36': source_sheet.cell(row=29, column=20).value,
                    'B36': source_sheet.cell(row=30, column=20).value,
                    'A37': source_sheet.cell(row=29, column=27).value,
                    'B37': source_sheet.cell(row=30, column=27).value,
                    'A39': source_sheet.cell(row=54, column=2).value,
                    'A40': source_sheet.cell(row=55, column=2).value,
                    'B40': source_sheet.cell(row=56, column=2).value,
                    'A41': source_sheet.cell(row=55, column=8).value,
                    'B41': source_sheet.cell(row=56, column=8).value,
                    'A42': source_sheet.cell(row=55, column=14).value,
                    'B42': source_sheet.cell(row=56, column=14).value,
                    'A43': source_sheet.cell(row=55, column=25).value,
                    'B43': source_sheet.cell(row=56, column=25).value,
                    'A44': source_sheet.cell(row=55, column=29).value,
                    'B44': source_sheet.cell(row=56, column=29).value,
                    'A46': source_sheet.cell(row=58, column=2).value,
                    'A47': source_sheet.cell(row=59, column=2).value,
                    'B47': source_sheet.cell(row=60, column=2).value,
                    'A48': source_sheet.cell(row=59, column=8).value,
                    'B48': source_sheet.cell(row=60, column=8).value,
                    'A49': source_sheet.cell(row=59, column=11).value,
                    'B49': source_sheet.cell(row=60, column=11).value,
                    'A50': source_sheet.cell(row=59, column=15).value,
                    'B50': source_sheet.cell(row=60, column=15).value,
                    'A51': source_sheet.cell(row=59, column=19).value,
                    'B51': source_sheet.cell(row=60, column=19).value,
                    'A52': source_sheet.cell(row=59, column=37).value,
                    'B52': source_sheet.cell(row=60, column=37).value,
                    # 'A58': source_sheet.cell(row=83, column=2).value,
                    # 'A59': source_sheet.cell(row=85, column=2).value,
                    # 'B59': source_sheet.cell(row=86, column=2).value,
                    # 'A60': source_sheet.cell(row=85, column=30).value,
                    # 'B60': source_sheet.cell(row=86, column=30).value,
                    'A54': source_sheet.cell(row=88, column=2).value,
                    'A55': source_sheet.cell(row=89, column=2).value,
                    'B55': source_sheet.cell(row=90, column=2).value,
                    'A56': source_sheet.cell(row=89, column=9).value,
                    'B56': source_sheet.cell(row=90, column=9).value,
                    'A57': source_sheet.cell(row=89, column=14).value,
                    'B57': source_sheet.cell(row=90, column=14).value,
                    'A58': source_sheet.cell(row=89, column=21).value,
                    'B58': source_sheet.cell(row=90, column=21).value,
                    'A59': source_sheet.cell(row=89, column=43).value,
                    'B59': source_sheet.cell(row=90, column=43).value,
                }

                data_to_append2 = {
                    'G1': f"ОСОБА 2",
                    'G2': source_sheet2.cell(row=6, column=2).value,
                    'H2': source_sheet2.cell(row=7, column=2).value,
                    'G3': source_sheet2.cell(row=6, column=8).value,
                    'H3': source_sheet2.cell(row=7, column=8).value,
                    'G4': source_sheet2.cell(row=6, column=14).value,
                    'H4': source_sheet2.cell(row=7, column=14).value,
                    'G5': source_sheet2.cell(row=6, column=24).value,
                    'H5': source_sheet2.cell(row=7, column=24).value,
                    'G6': source_sheet2.cell(row=6, column=28).value,
                    'H6': source_sheet2.cell(row=7, column=28).value,
                    'G9': source_sheet2.cell(row=12, column=2).value,
                    'G10': source_sheet2.cell(row=13, column=2).value,
                    'H10': source_sheet2.cell(row=14, column=2).value,
                    'G11': source_sheet2.cell(row=13, column=8).value,
                    'H11': source_sheet2.cell(row=14, column=8).value,
                    'G12': source_sheet2.cell(row=13, column=11).value,
                    'H12': source_sheet2.cell(row=14, column=11).value,
                    'G13': source_sheet2.cell(row=13, column=15).value,
                    'H13': source_sheet2.cell(row=14, column=15).value,
                    'G14': source_sheet2.cell(row=13, column=19).value,
                    'H14': source_sheet2.cell(row=14, column=19).value,
                }

                combined_data = {**data_to_append1, **data_to_append2}

                for cell, value in combined_data.items():
                    destination_sheet[cell] = value
                    if cell.startswith('B') or cell.startswith('H'):
                        destination_sheet[cell].alignment = left_alignment  # Apply left alignment to cell in column B

                # Format cell B5 as a date cell and apply left alignment
                date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY', alignment=left_alignment)
                destination_sheet['B5'].style = date_style
                destination_sheet['H5'].style = date_style

                # Save the destination workbook
                cleaned_file_name = re.sub(r'[<>:"/\|?*]', '', file_name_value)[:30]  # Limit to 30 characters
                destination_file = f'{cleaned_file_name}.xlsx'
                destination_workbook.save(os.path.join(output_folder, destination_file))

                # Close both workbooks
                source_workbook.close()
                destination_workbook.close()

                processed_files += 1

                current_time = time.time()
                if current_time - last_update_time >= update_interval:
                    progress = (processed_files / total_files) * 100
                    print(f"Progress: {progress:.2f}% ({processed_files}/{total_files} files)")
                    last_update_time = current_time

            except Exception as e:
                print(f"Error processing {source_file}: {str(e)}")

# Stop the timer
end_time = time.time()
elapsed_time = end_time - start_time
print(f"Finished in {elapsed_time:.2f} seconds")

